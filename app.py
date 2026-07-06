import os
import requests
from flask import Flask, jsonify, render_template, request
from openpyxl import load_workbook
from pathlib import Path
from dotenv import load_dotenv

load_dotenv(Path(__file__).parent / ".env")

app = Flask(__name__)

TMDB_API_KEY = os.getenv("TMDB_API_KEY")
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")

# In-memory poster cache
poster_cache = {}
overview_cache = {}
actors_cache = {}
providers_cache = {}

# ---------------------------------------------------------------------------
# Load movies from xlsx
# ---------------------------------------------------------------------------

def load_movies():
    xlsx_path = os.path.join(os.path.dirname(__file__), "bbedca29-d25d-47ae-add7-e6680046ff89.xlsx")
    wb = load_workbook(xlsx_path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    movies = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[1]:  # skip rows without IMDb ID
            continue

        movie = {}
        for i, header in enumerate(headers):
            movie[header] = row[i]

        # Normalise dates to ISO strings
        for date_key in ("Created", "Modified", "Release Date", "Date Rated"):
            val = movie.get(date_key)
            if hasattr(val, "strftime"):
                movie[date_key] = val.strftime("%Y-%m-%d")
            else:
                movie[date_key] = None

        # Fix IMDb Rating: stored as integer ×10 (e.g. 87 → 8.7)
        imdb_raw = movie.get("IMDb Rating")
        if isinstance(imdb_raw, int):
            movie["IMDb Rating"] = imdb_raw / 10

        # Split genres into a list
        genres_raw = movie.get("Genres") or ""
        movie["genres_list"] = [g.strip() for g in genres_raw.split(",") if g.strip()]

        movies.append(movie)

    return movies


MOVIES = load_movies()

# Pre-build a compact summary for the chat system prompt (avoid sending huge JSON)
def build_movie_summary():
    lines = []
    for m in MOVIES:
        rating = m.get("Your Rating")
        rating_str = f" | Your rating: {rating}/10" if rating else " | Not rated yet"
        lines.append(
            f"- {m['Title']} ({m['Year']}) — Dir: {m['Directors']} | IMDb: {m['IMDb Rating']}{rating_str} | Genres: {m['Genres']}"
        )
    return "\n".join(lines)

MOVIE_SUMMARY = build_movie_summary()


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/movies")
def api_movies():
    return jsonify(MOVIES)


@app.route("/api/movies/<imdb_id>/poster")
def api_poster(imdb_id):
    if imdb_id in poster_cache:
        return jsonify({
            "poster": poster_cache[imdb_id], 
            "overview": overview_cache.get(imdb_id),
            "actors": actors_cache.get(imdb_id) or [],
            "providers": providers_cache.get(imdb_id) or []
        })

    if not TMDB_API_KEY:
        poster_cache[imdb_id] = None
        overview_cache[imdb_id] = None
        actors_cache[imdb_id] = []
        providers_cache[imdb_id] = []
        return jsonify({"poster": None, "overview": None, "actors": [], "providers": []})

    try:
        url = f"https://api.themoviedb.org/3/find/{imdb_id}"
        params = {"external_source": "imdb_id", "api_key": TMDB_API_KEY}
        resp = requests.get(url, params=params, timeout=8)
        data = resp.json()
        results = data.get("movie_results", [])
        is_tv = False
        if not results:
            results = data.get("tv_results", [])
            is_tv = True

        if results:
            r = results[0]
            poster_cache[imdb_id] = f"https://image.tmdb.org/t/p/w300{r['poster_path']}" if r.get("poster_path") else None
            overview_cache[imdb_id] = r.get("overview") or None
            
            # Fetch cast/credits and watch providers using TMDB ID
            actors = []
            providers = []
            tmdb_id = r.get("id")
            if tmdb_id:
                try:
                    media_type = "tv" if is_tv else "movie"
                    details_url = f"https://api.themoviedb.org/3/{media_type}/{tmdb_id}"
                    details_params = {"api_key": TMDB_API_KEY, "append_to_response": "credits,watch/providers"}
                    details_resp = requests.get(details_url, params=details_params, timeout=5)
                    details_data = details_resp.json()
                    
                    credits_key = "aggregate_credits" if is_tv and "aggregate_credits" in details_data else "credits"
                    cast = details_data.get(credits_key, {}).get("cast", [])
                    actors = [member.get("name") for member in cast[:5] if member.get("name")]
                    
                    # Parse watch providers for Spain (ES)
                    wp_data = details_data.get("watch/providers", {})
                    es_data = wp_data.get("results", {}).get("ES", {})
                    flatrate = es_data.get("flatrate", [])
                    for p in flatrate:
                        p_name = p.get("provider_name")
                        logo_path = p.get("logo_path")
                        if p_name:
                            providers.append({
                                "name": p_name,
                                "logo": f"https://image.tmdb.org/t/p/w92{logo_path}" if logo_path else None
                            })
                except Exception:
                    pass
            actors_cache[imdb_id] = actors
            providers_cache[imdb_id] = providers
        else:
            poster_cache[imdb_id] = None
            overview_cache[imdb_id] = None
            actors_cache[imdb_id] = []
            providers_cache[imdb_id] = []
    except Exception:
        poster_cache[imdb_id] = None
        overview_cache[imdb_id] = None
        actors_cache[imdb_id] = []
        providers_cache[imdb_id] = []

    return jsonify({
        "poster": poster_cache[imdb_id], 
        "overview": overview_cache.get(imdb_id),
        "actors": actors_cache.get(imdb_id) or [],
        "providers": providers_cache.get(imdb_id) or []
    })


@app.route("/api/chat", methods=["POST"])
def api_chat():
    if not OPENAI_API_KEY:
        return jsonify({
            "reply": (
                "⚠️ No OpenAI API key found. Add OPENAI_API_KEY=your_key to your .env file "
                "and restart the app to enable the AI assistant."
            )
        })

    try:
        from openai import OpenAI

        body = request.get_json(force=True)
        user_message = body.get("message", "").strip()
        history = body.get("history", [])

        if not user_message:
            return jsonify({"reply": "Please type a message."})

        system_prompt = (
            "You are a friendly film enthusiast assistant for a personal movie collection app "
            "with retro Blockbuster vibes. You have access to the user's complete movie collection below.\n\n"
            "MOVIE COLLECTION:\n"
            f"{MOVIE_SUMMARY}\n\n"
            "Help the user discover movies from their list, discuss films, give recommendations based on "
            "what they've rated highly, compare movies, and chat about cinema in general. "
            "Keep responses concise and engaging. Use the collection data to give personalised answers. "
            "You can reference their personal ratings when relevant."
        )

        messages = [{"role": "system", "content": system_prompt}]
        for h in history:
            role = h.get("role")
            content = h.get("content", "")
            if role in ("user", "assistant") and content:
                messages.append({"role": role, "content": content})
        messages.append({"role": "user", "content": user_message})

        client = OpenAI(api_key=OPENAI_API_KEY)
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=messages,
            max_tokens=1024,
        )

        reply = response.choices[0].message.content

        # Save to chat_history.json
        try:
            import json
            from datetime import datetime
            history_file = os.path.join(os.path.dirname(__file__), "chat_history.json")
            
            if os.path.exists(history_file):
                with open(history_file, "r", encoding="utf-8") as f:
                    try:
                        history_data = json.load(f)
                        if not isinstance(history_data, list):
                            history_data = []
                    except Exception:
                        history_data = []
            else:
                history_data = []
                
            timestamp = datetime.now().isoformat()
            history_data.append({
                "timestamp": timestamp,
                "role": "user",
                "content": user_message
            })
            history_data.append({
                "timestamp": timestamp,
                "role": "assistant",
                "content": reply
            })
            
            with open(history_file, "w", encoding="utf-8") as f:
                json.dump(history_data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"Error saving chat history: {e}")

        return jsonify({"reply": reply})

    except Exception as e:
        return jsonify({"reply": f"Error communicating with AI: {str(e)}"})


@app.route("/api/chat/history")
def api_chat_history():
    import json
    history_file = os.path.join(os.path.dirname(__file__), "chat_history.json")
    if os.path.exists(history_file):
        try:
            with open(history_file, "r", encoding="utf-8") as f:
                history_data = json.load(f)
                if isinstance(history_data, list):
                    # Return last 40 messages to avoid sending too much data
                    return jsonify(history_data[-40:])
        except Exception:
            pass
    return jsonify([])


if __name__ == "__main__":
    app.run(debug=True, port=5000)
